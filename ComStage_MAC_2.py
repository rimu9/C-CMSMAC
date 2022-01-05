# -*- coding: utf-8 -*-
import copy
import gc
import random
import math
import matplotlib.pyplot as plt
import networkx as net
import logging
from openpyxl import load_workbook
import common.settings as settings
from Communication.simulation.ComObjectCollection import *
from Communication.simulation.ComStage import ComStage
from Communication.simulation.ComRobot import ComRobot
from Communication.simulation.ComRobotAF_MAC_2 import ComRobotAFMAC2, VLCMsg


class ComStageMAC2(ComStage):
    nowTime = 0  # 当前时间点记录

    def __init__(self):
        super().__init__()
        self.formerSlots = 20  # 一帧多少正式槽
        self.preSlots = 5  # 一帧多少预备槽

        self.rbtSlotsID = []  # 这个数组里面，为每个slot定义一个set()，里面存放在该时刻发消息的机器人id
        for i in range(self.formerSlots + self.preSlots):
            self.rbtSlotsID.append(set())

        self.collRobots = []  # 全局存放存在冲突的连接，[(发出信号的机器人id， 接收信号的机器人id)，...]

        # 指标
        self.accessCountHistory = []  # 储存每个机器人接入时间
        self.clusters = []  # 储存聚集系数
        self.shortestPathLen = []  # 储存平均最短路径
        self.overloadCount = 0  # 储存随着节点的增多，整体的接入负载。
        self.overloadHistory = []
        self.collPerFrame = 0  # 计算每个frame中的冲突有多少个。
        self.collPerFrameHistory = []
        self.fairnessCounter = []  # 计算fairness用。
        self.fairnessRecoder = []  # 记录fairness。

        # 标志位
        self.acesSuccessFlag = 1  # 机器人成功接入网络，再生成新的机器人进来。
        self.endLoopFlag = 0  # 结束一次循环，从头重新开始仿真标志位
        self.writexlsxFlag = 0  # 往Excel表格中写数据
        self.calcMetricsFlag = 0  # 置1计算一次指标。

        # 以下变量就是构造仿真地图所用，对协议本身没啥关系，主要是为了方便保证机器人能够构造出一个连通图出来
        self.lineNum = 12
        self.scale = self.mEnvSize[0]
        self.map = []
        self.map = [[] for i in range(self.lineNum - 2)]
        for i in range(1, self.lineNum - 1):
            for j in range(1, self.lineNum - 1):
                x = i / self.lineNum * self.scale - self.scale / 2
                y = j / self.lineNum * self.scale - self.scale / 2
                self.map[i - 1].append({'status': 'idle',  # idle空闲，busy占用, useless该位置无法接入
                                        'pos': (x, y)})

    # region 与地图map相关函数
    '''
    description: 设置self.map中的某个点的状态为busy
    param: i，j就是点在self.map中的位置的索引值
    return: none
    '''
    def setMapBusy(self, i, j):
        self.map[i][j]['status'] = 'busy'

    '''
    description: 设置self.map中的某个点的状态为idle
    param: i，j就是点在self.map中的位置的索引值
    return: none
    '''
    def setMapIdle(self, i, j):
        self.map[i][j]['status'] = 'idle'

    '''
    description: 设置self.map中的某个点的状态为useless
    param: i，j就是点在self.map中的位置的索引值
    return: none
    '''
    def setMapUseless(self, i, j):
        self.map[i][j]['status'] = 'useless'

    '''
    description: 清除map中的useless点。因为本轮机器人已经接入完成了，下一轮机器人再接入的时候，原本useless的位置可能这一轮就可以用了。
    param: none
    return: none
    '''
    def refreshMap(self):
        for i in range(self.lineNum - 2):
            for j in range(self.lineNum - 2):
                if self.map[i][j]['status'] == 'useless':
                    self.setMapIdle(i, j)
                    print('%%%%%%%%%%%%%%%%%%%%%%%%%%%%% 清除一个无用的位置', i, j, '%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%')

    '''
    description: 判断第map[i][j]点的前后左右（四周）几个点是否有是被占用的。因为updateMap函数中，candidates中包含所有idle的点，其中有很多不切实际的位置，筛选一部分掉。
    param: i，j就是点在self.map中的位置的索引值
    return: 0 or 1
    '''
    def verifyPos(self, i, j):
        if self.map[max(0, i - 1)][j]['status'] == 'busy' or \
           self.map[min(i + 1, 9)][j]['status'] == 'busy' or \
           self.map[i][max(0, j - 1)]['status'] == 'busy' or \
           self.map[i][min(j + 1, 9)]['status'] == 'busy' or \
           self.map[min(i + 1, 9)][min(j + 1, 9)]['status'] == 'busy' or \
           self.map[max(0, i - 1)][max(0, j - 1)]['status'] == 'busy' or \
           self.map[min(i + 1, 9)][max(0, j - 1)]['status'] == 'busy' or \
           self.map[max(0, i - 1)][min(j + 1, 9)]['status'] == 'busy':
            return 1
        else:
            return 0

    '''
    description: 从self.map中返回一个空闲的位置，坐标，作为机器人的目标位置
    param: none
    return:  pos index in map,     pos
             tuple,          ,     np.array([x,y,z], float32)
             calc from self.map    calc from self.map
    '''
    def updateMap(self):
        candidates = []
        for i in range(5):
            mi = 4 - i  # max mi=4  min mi=0
            ma = 5 + i  # max ma=9  min ma=5
            for p in range(mi, ma + 1):
                if self.map[mi][p]['status'] == 'idle' and self.verifyPos(mi, p):
                    candidates.append((mi, p))
                if self.map[ma][p]['status'] == 'idle' and self.verifyPos(ma, p):
                    candidates.append((ma, p))
                if self.map[p][mi]['status'] == 'idle' and self.verifyPos(p, mi):
                    candidates.append((p, mi))
                if self.map[p][ma]['status'] == 'idle' and self.verifyPos(p, ma):
                    candidates.append((p, ma))

        if len(candidates) < 1:
            # 遍历了所有的选择都不能找到一个好的位置接入，则随机找一个位置接入
            range_x = (-900, 900)
            range_y = (-900, 900)
            x = random.uniform(range_x[0], range_x[1])
            y = random.uniform(range_y[0], range_y[1])
            pos = np.array([x, y, 500], dtype=np.float32)

            print('stage.map 分配一个随机的位置给新机器人 ---------------------------------------------------------')

            return (-1, -1), pos

        posIndex = random.sample(candidates, 1)[0]
        pos = np.array([self.map[posIndex[0]][posIndex[1]]['pos'][0],
                        self.map[posIndex[0]][posIndex[1]]['pos'][1],
                        500], dtype=np.float32)

        self.setMapBusy(posIndex[0], posIndex[1])

        return posIndex, pos

    # endregion函数

    '''
    description: 从全局的角度来看机器人网络，返回从全局观测到的机器人robotId的入邻居。因为机器人自身观测到的入邻居可能是不完整的，由于冲突的存在。
                 这一步主要就是为了方便保证将新机器人融入到网络中形成一个连通图。没有实际意义。
    param: checkId: 需要检查的网络中的哪一个机器人的id
    return:  globalIn: 入邻居对象数组
    '''
    def calcRobotIn(self, checkId):
        ret = []
        for robot in self.mRobotList.nodes:
            outIds = [rbt.mId for rbt in robot.outRobots]
            if checkId in outIds:
                ret.append(robot)
        return ret

    '''
    description: 将指标写入Excel表格中，记得每轮仿真前确认对应路径中文件是存在的。
    param: none
    return:  none
    '''
    def writeLogFile(self):
        file1Name = "../../resultHJ/accessTime.xlsx"
        file1 = load_workbook(file1Name)
        sheet1 = file1[file1.sheetnames[0]]
        max1 = sheet1.max_row + 1
        l = 1
        for item in self.accessCountHistory:
            sheet1.cell(max1, l, str(item))
            l += 1
        file1.save(file1Name)

        file2Name = "../../resultHJ/clusters.xlsx"
        file2 = load_workbook(file2Name)
        sheet2 = file2[file2.sheetnames[0]]
        max2 = sheet2.max_row + 1
        l = 1
        for item in self.clusters:
            sheet2.cell(max2, l, str(item))
            l += 1
        file2.save(file2Name)

        file3Name = "../../resultHJ/shortestpath.xlsx"
        file3 = load_workbook(file3Name)
        sheet3 = file3[file3.sheetnames[0]]
        max3 = sheet3.max_row + 1
        l = 1
        for item in self.shortestPathLen:
            sheet3.cell(max3, l, str(item))
            l += 1
        file3.save(file3Name)

        file4Name = "../../resultHJ/overload.xlsx"
        file4 = load_workbook(file4Name)
        sheet4 = file4[file4.sheetnames[0]]
        max4 = sheet4.max_row + 1
        l = 1
        for item in self.overloadHistory:
            sheet4.cell(max4, l, str(item))
            l += 1
        file4.save(file4Name)

        file5Name = "../../resultHJ/collperframe.xlsx"
        file5 = load_workbook(file5Name)
        sheet5 = file5[file5.sheetnames[0]]
        r = sheet5.max_row + 1
        c = 1
        for item in self.collPerFrameHistory:
            sheet5.cell(r, c, str(item))
            c += 1
            if c == 100:
                c = 1
                r += 1
        file5.save(file5Name)

        file6Name = "../../resultHJ/fairness.xlsx"
        file6 = load_workbook(file6Name)
        sheet6 = file6[file6.sheetnames[0]]
        r = sheet6.max_row + 1
        c = 1
        for item in self.fairnessRecoder:
            sheet6.cell(r, c, str(item))
            c += 1
        file6.save(file6Name)

        # print('ACCESS TIME: ', self.accessCountHistory)
        # print('CLUSTERS: ', self.clusters)
        # print('SHORTEST PATH LEN: ', self.shortestPathLen)
        # print('OVERLOAD: ', self.overloadHistory)
        # print('COLLISION PER FRAME: ', self.collPerFrameHistory)

    '''
    description: 计算指标，存入数组中。
    param: none
    return:  none
    '''
    def calcNetworkMetrics(self):
        numr = len(self.mRobotList.nodes)

        # cluster = net.average_clustering(self.mRobotList)
        # self.clusters.append((num, cluster))
        # # print('robots num: ', num, ' with clusters: ', cluster)
        #
        # shortPath = net.average_shortest_path_length(self.mRobotList)
        # self.shortestPathLen.append((num, shortPath))
        # # print('robots num: ', num, ' with shortest path: ', shortpath)

        self.overloadHistory.append((numr, self.overloadCount))

        for robot in self.mRobotList.nodes:
            if robot.status != 'online':
                raise Exception('ERROR(in calcNetworkMetrics): cannot calc Fairness metric cause there exist a not online robot')
        num = 0
        den = 0
        for count in self.fairnessCounter:
            num += count
            den += math.pow(count, 2)
        if den == 0:
            if len(self.mRobotList.nodes) != 5:
                raise Exception('ERROR(in calcNetworkMetrics): cannot calc Fairness metric cause den equals to 0')
        try:
            fairness = math.pow(num, 2) / (numr * den)
            self.fairnessRecoder.append((numr, fairness))
        except ZeroDivisionError:
            pass

        self.fairnessCounter.clear()
        self.fairnessCounter = [0 for i in range(numr)]

    '''
    description: 生成新的机器人对象，添加到网络中。
    param: addType: 'signal'每次生成一个机器人 or 'multi'每次生成多个机器人
    return:  num: 生成的机器人的数量
    '''
    def addNewRobots(self, addType):
        if addType == 'signal':
            num = 1
            range_x = (-stage.mEnvSize[0], stage.mEnvSize[0])
            range_y = (-stage.mEnvSize[1], stage.mEnvSize[1])
            x = random.uniform(range_x[0], range_x[1])
            y = random.uniform(range_y[0], range_y[1])
            newrobot = ComRobotAFMAC2((x, y, 400), self)
            newrobot.isPlotTrail = True
            newrobot.setRadius(2)
            stage.addRobot(newrobot)
            print('**********a new robot has generate ---id:', newrobot.mId, '*****************')
        else:
            num = random.randint(2, 4)
            for i in range(num):
                range_x = (-stage.mEnvSize[0], stage.mEnvSize[0])
                range_y = (-stage.mEnvSize[1], stage.mEnvSize[1])
                x = random.uniform(range_x[0], range_x[1])
                y = random.uniform(range_y[0], range_y[1])
                newrobot = ComRobotAFMAC2((x, y, 400), self)
                newrobot.isPlotTrail = True
                newrobot.setRadius(2)
                stage.addRobot(newrobot)
                print('**********a new robot has generate ---id:', newrobot.mId, '*****************')
        return num

    '''
    description: 站在全局的角度取判断当前时间下是否存在冲突。有冲突换成红色线条，没有冲突的就是蓝色线条，绘制成图
    param: none
    return: edgesColor： list, 每条边对应的颜色数组，用于后期画图
    '''
    def collectCollisions(self):
        edgesColor = []

        self.mRobotList.clear_edges()
        self.collRobots.clear()

        senders = list(self.rbtSlotsID[ComStageMAC2.nowTime])
        notsenders = list(set(range(ComRobot._robot_count)) - set(senders))

        # 先把不发送信息的边画到图上，白色。
        for id in notsenders:
            robot = list(self.mRobotList.nodes)[id]
            for r in robot.outRobots:
                self.mRobotList.add_edge(robot, r)
                edgesColor.append('w')

        # 下面再分析发送信息的机器人之间有没有冲突
        if len(senders) < 2:
            for sender in senders:
                robot = list(self.mRobotList.nodes)[sender]

                logging.info('\tsneder id: ' + str(robot.mId) + ' outers: ' + str([r.mId for r in robot.outRobots])[1: -1])

                for r in robot.outRobots:
                    self.mRobotList.add_edge(robot, r)
                    edgesColor.append('b')

                    if robot.status == 'online':
                        try:
                            self.fairnessCounter[sender] += 1
                        except IndexError:
                            if len(self.mRobotList.nodes) != 5:
                                raise Exception('ERROR(in collectCollisions - add fairness): IndexError')

            return edgesColor

        for i in range(len(senders) - 1):
            roboti = list(self.mRobotList.nodes)[senders[i]]
            oi = {r.mId for r in roboti.outRobots if r.status != 'entering'}

            for j in range(i + 1, len(senders)):
                robotj = list(self.mRobotList.nodes)[senders[j]]
                oj = {r.mId for r in robotj.outRobots if r.status != 'entering'}
                cos = oi & oj
                if cos:
                    for co in cos:
                        self.collRobots.extend([(senders[i], co), (senders[j], co)])

                    logging.info('\tcollision happened: ' + str(senders[i]) + ' and ' + str(senders[j]) + ' --> ' + str(cos)[1: -1])

                    # 冲突发生，计数。
                    self.collPerFrame += 2 * len(cos)

        for sender in senders:
            robot = list(self.mRobotList.nodes)[sender]

            logging.info('\tsneder id: ' + str(robot.mId) + ' outers: ' + str([r.mId for r in robot.outRobots])[1: -1])

            for r in robot.outRobots:
                self.mRobotList.add_edge(robot, r)
                if (robot.mId, r.mId) in self.collRobots:
                    edgesColor.append('r')
                else:
                    edgesColor.append('b')

                    if robot.status == 'online':
                        try:
                            self.fairnessCounter[sender] += 1
                        except IndexError:
                            if len(self.mRobotList.nodes) != 5:
                                raise Exception('ERROR(in collectCollisions - add fairness): IndexError')

        return edgesColor

    def update(self):
        # 图坐标系建立
        if self.mAx:  # 画3d图的准备
            self.mAx.cla()
            self.mAx.grid(False)
            self.mAx.set_xlim(-self.mEnvSize[0], self.mEnvSize[0])
            self.mAx.set_ylim(-self.mEnvSize[1], self.mEnvSize[1])
            self.mAx.set_zlim(-self.mEnvSize[2], self.mEnvSize[2])
        if self.mGraphAx:  # 画连通图的准备
            self.mGraphAx.cla()

        # 此处循环是为了感知周围的邻居，获得出邻居们
        for robot in stage.mRobotList.nodes:
            robot.sense()

        # 执行机器人的碰撞规避操作
        logging.info('Stage Info: (func collsSolve):')
        tempSet = set()
        for robot in self.mRobotList.nodes:
            tempSet |= robot.collsSolve()
        logging.info('\tEnd')

        # # 梳理网络结构，确定有向连通图
        logging.info('Stage Info: (func collectCollisions):')
        edgescolor = self.collectCollisions()
        logging.info('\tEnd')

        # 执行信息交互。
        logging.info('Robot Info: (func communicateTo):')
        for robot in self.mRobotList.nodes:
            self.overloadCount += robot.communicateTo(ComStageMAC2.nowTime)
        logging.info('\tEnd')

        # 执行完通信操作以后，就把tempSet中的数据重新加回来，因为每次机器人接收的信息可能不同步，不是每次判断都是对的，所以采取的措施就是我每次根据互斥信息进行判断，而不是判断一次就永久删除。
        for item in tempSet:
            robot = list(self.mRobotList.nodes)[item[0]]
            if robot.mId != item[0]:
                print('\033[1;30;31m\tERROR: robot searching error\033[0m')
                logging.info('ERROR: robot searching error in add collision slots into the original robot')
                return
            robot.addSlots({item[1]})

        # 每个机器人对接收的信息进行处理
        logging.info('Robot Info: (func updatePre):')
        for robot in stage.mRobotList.nodes:
            robot.updatePre()
        logging.info('\tEnd')

        # 每个机器人执行更新操作
        logging.info('Robot Info: (func update):')
        for robot in stage.mRobotList.nodes:
            robot.update()
            if self.mAx:
                robot.draw(self.mAx)
        logging.info('\tEnd')

        # 图形显示
        if self.isPlotGraph and len(self.mRobotList.nodes) > 1:
            graph_pos = net.circular_layout(self.mRobotList)  # 布置框架
            net.draw(self.mRobotList, graph_pos, ax=self.mGraphAx,
                     with_labels=False, node_size=30,
                     edge_color=edgescolor)

        # 更新时间点
        ComStageMAC2.nowTime += 1
        if ComStageMAC2.nowTime >= (self.formerSlots + self.preSlots):
            ComStageMAC2.nowTime = 0
            # 1frame结束了，清零冲突计数器
            self.collPerFrameHistory.append(self.collPerFrame)
            self.collPerFrame = 0

        # 计算指标
        if self.calcMetricsFlag == 1:
            self.calcNetworkMetrics()
            self.calcMetricsFlag = 0

        # 如果所有新机器人都接入完成了，就计算指标，刷新地图，生成新的机器人。
        if self.acesSuccessFlag == 0:
            # 清除map中设置为无用的点
            self.refreshMap()
            # 新生成一个机器人或者多个机器人
            self.acesSuccessFlag = self.addNewRobots('signal')

        # 设置当网络有100个机器人的时候，就停止仿真，毕竟是小尺寸网络， 后期可以试试大一点的网络。
        if len(self.mRobotList) >= 150 and self.writexlsxFlag == 0:
            self.writeLogFile()
            self.writexlsxFlag = 1
            self.endLoopFlag = 1

    def run(self):
        # self.initEnv()
        while True:
            if self.count % self.mSavePosRound == 0:
                if self.isSavePos2:
                    self.isSavePos = True
            else:
                self.isSavePos = False
            if (self.count * settings.CS_INTERVAL) >= self.mRuningTime:
                break
            self.count += 1
            if self.count % 1 == 0:
                print("\nRound: 【%d : %d】-START-" % (self.count, self.mRuningTime/settings.CS_INTERVAL),
                      'now time:', ComStageMAC2.nowTime)
                logging.info("Round: 【%d : %d】-START-" % (self.count, self.mRuningTime/settings.CS_INTERVAL) +
                             '  now time:%d' % ComStageMAC2.nowTime +
                             '  senders: ' + str(stage.rbtSlotsID[ComStageMAC2.nowTime]))
            self.update()
            # print("     -END-\n")
            logging.info('END\n ')
            plt.pause(settings.CS_INTERVAL)

            if self.endLoopFlag == 1:
                print('程序运行一次退出')
                break

    def initEnv(self):
        """
        初始化环境
        :return:
        """
        if self.mFig is None:
            self.mFig = plt.figure(figsize=self.mFigSize, constrained_layout=True)
        if self.isPlotGraph:
            gs = self.mFig.add_gridspec(8, 15)
            self.mAx = self.mFig.add_subplot(gs[:, 0:7], projection="3d")
            self.mGraphAx = self.mFig.add_subplot(gs[:, 10:15])

        else:
            self.mAx = self.mFig.add_subplot(projection="3d")

        for surf in self.mSurfaceGroup:
            surf.setAx(self.mAx)


if __name__ == "__main__":
    # region unit test

    # region test robot.updateInRobots()函数
    # stage = ComStageMAC2()
    # robot0 = ComRobotAFMAC2((stage.map[4][4]['pos'][0], stage.map[4][4]['pos'][1], 500), stage)
    # stage.addRobot(robot0)
    # for i in range(30):
    #     robot = ComRobotAFMAC2((stage.map[4][4]['pos'][0], stage.map[4][4]['pos'][1], 500), stage)
    #     stage.addRobot(robot)
    # constants = list(range(20))
    # for i in range(50):
    #     tempIn = random.sample(constants, 3)
    #     robot0.updateInRobots(tempIn)
    #     print(stage.nowTime)
    #     print(i, ' tempIn: ', tempIn)
    #     print(i, ' inRobotsIdsRecoder: ', len(robot0.inRobotsIdsRecoder), robot0.inRobotsIdsRecoder)
    #     ComStageMAC2.nowTime += 1
    #     if ComStageMAC2.nowTime >= (stage.formerSlots + stage.preSlots):
    #         ComStageMAC2.nowTime = 0
    # endregion

    # region test robot.processRecvMsg()函数
    # from Communication.simulation.ComRobotAF_MAC import VLCMsg
    # stage = ComStageMAC2()
    # robot0 = ComRobotAFMAC2((stage.map[4][4]['pos'][0], stage.map[4][4]['pos'][1], 500), stage)
    # robot0.status = 'online'
    # stage.addRobot(robot0)
    #
    # msg1 = VLCMsg(4, isFlooding=True)
    # msg1.seq = 5
    # msg1.payload['REQ'] = [1,2,3]
    # robot0.recvMsgBuffer.append(msg1)
    #
    # msg1 = VLCMsg(4, isFlooding=True)
    # msg1.seq = 1
    # msg1.payload['APP'] = None
    # robot0.recvMsgBuffer.append(msg1)
    #
    # msg1 = VLCMsg(4, isFlooding=True)
    # msg1.seq = 3
    # msg1.payload['NEIGHBORS'] = [[4, 11, 11, 12, 13], [4, 5, 5, 7]]
    # robot0.recvMsgBuffer.append(msg1)
    #
    # msg1 = VLCMsg(4, isFlooding=True)
    # msg1.seq = 0
    # msg1.payload[9] = {'type': 'slot-add', 'slots': {11}}
    # robot0.recvMsgBuffer.append(msg1)
    #
    # msg1 = VLCMsg(4, isFlooding=True)
    # msg1.seq = 0
    # msg1.payload[9] = {'type': 'slot-rdu', 'slots': {19, 2}}
    # robot0.recvMsgBuffer.append(msg1)
    #
    # msg1 = VLCMsg(4, isFlooding=True)
    # msg1.seq = 0
    # msg1.payload[9] = {'type': 'slot-rpl', 'slots': {10, 19}}
    # robot0.recvMsgBuffer.append(msg1)
    #
    # msg1 = VLCMsg(4, isFlooding=True)
    # msg1.seq = 7
    # msg1.payload['REQ'] = [1, 2, 3]
    # robot0.recvMsgBuffer.append(msg1)
    #
    # msg1 = VLCMsg(4, isFlooding=True)
    # msg1.seq = 6
    # msg1.payload['NEIGHBORS'] = [[12, 13], [4, 5, 7, 8]]
    # robot0.recvMsgBuffer.append(msg1)
    #
    # ret = robot0.processRecvMsg()
    # print(ret)
    # endregion

    # region test robot.addCollisions()函数
    # stage = ComStageMAC2()
    # robot0 = ComRobotAFMAC2((stage.map[4][4]['pos'][0], stage.map[4][4]['pos'][1], 500), stage)
    # stage.addRobot(robot0)
    # robot0.addCollisions([{1, 2}, {3,4}, {2,3}, {1, 2},{1, 2, 3},{1, 2},{1, 3},{1, 2}, {1, 2}])
    # endregion

    # endregion

    # 打印日志
    logging.basicConfig(filename='../../resultHJ/logging2.log', level=logging.DEBUG)

    # 每一轮仿真开始前需要把上一轮仿真留下来的所有变量，状态全部初始化或者删除掉，gc和del就是干这个事的
    initdir = copy.deepcopy(list(globals().keys()))
    for i in range(10):
        for key in list(globals().keys()):
            if (key not in initdir) and (key != "key") and (key != "initdir") and (key != "i"):
                globals().pop(key)
        #  删除变量
        del key
        object_collection.clear()
        gc.collect()

        # 初始化类内置属性。
        ComRobot._robot_count = 0
        ComStage.mCount = 0
        ComStageMAC2.nowTime = 0

        logging.info('------------------------------------START A NEW LOOP ------------------------------\n'
                     '------------------------------------START A NEW LOOP ------------------------------\n'
                     '------------------------------------START A NEW LOOP ------------------------------')

        stage = ComStageMAC2()

        # region 初始化4个机器人作为原始机器人
        robot0 = ComRobotAFMAC2((stage.map[4][4]['pos'][0], stage.map[4][4]['pos'][1], 500), stage)
        stage.map[4][4]['status'] = 'busy'
        robot0.setDirection(60 / 180 * math.pi)
        robot1 = ComRobotAFMAC2((stage.map[4][5]['pos'][0], stage.map[4][5]['pos'][1], 500), stage)
        stage.map[4][5]['status'] = 'busy'
        robot1.setDirection(-40 / 180 * math.pi)
        robot2 = ComRobotAFMAC2((stage.map[5][4]['pos'][0], stage.map[5][4]['pos'][1], 500), stage)
        stage.map[5][4]['status'] = 'busy'
        robot2.setDirection(135 / 180 * math.pi)
        robot3 = ComRobotAFMAC2((stage.map[5][5]['pos'][0], stage.map[5][5]['pos'][1], 500), stage)
        stage.map[5][5]['status'] = 'busy'
        robot3.setDirection(-100 / 180 * math.pi)
        robot0.slots.add(0)
        robot0.slots.add(1)
        robot1.slots.add(2)
        robot1.slots.add(3)
        robot2.slots.add(4)
        robot2.slots.add(5)
        robot3.slots.add(6)
        robot3.slots.add(7)
        stage.addRobot(robot0)
        stage.addRobot(robot1)
        stage.addRobot(robot2)
        stage.addRobot(robot3)

        robotlist = [robot0, robot1, robot2, robot3]
        for robot in robotlist:
            robot.setRadius(2)
            robot.status = 'online'
            robot.proSlotInfo[robot0.mId] = {'type': '', 'slots': copy.deepcopy(robot0.slots)}
            robot.proSlotInfo[robot1.mId] = {'type': '', 'slots': copy.deepcopy(robot1.slots)}
            robot.proSlotInfo[robot2.mId] = {'type': '', 'slots': copy.deepcopy(robot2.slots)}
            robot.proSlotInfo[robot3.mId] = {'type': '', 'slots': copy.deepcopy(robot3.slots)}
            for s in robot.slots:
                stage.rbtSlotsID[s].add(robot.mId)

            robot.graph.add_nodes_from([0, 1, 2, 3])
            robot.graph.add_edges_from([(0, 1), (0, 3), (1, 2), (1, 3), (2, 0), (2, 1), (2, 3), (3, 0), (3, 2)])

            robot.inRobotsIdsRecoder = [[] for i in range(21)]

        robot0.inRobots.extend([robot2, robot3])
        robot0.outRobots.extend([robot1, robot3])
        robot0.oldOutRobotsRecoder = {1: 'online', 3: 'online'}

        robot1.inRobots.extend([robot0, robot2])  # 0 2 3 ?
        robot1.outRobots.extend([robot2, robot3])
        robot1.oldOutRobotsRecoder = {2: 'online', 3: 'online'}

        robot2.inRobots.extend([robot1, robot3])
        robot2.outRobots.extend([robot0, robot1, robot3])
        robot2.oldOutRobotsRecoder = {0: 'online', 1: 'online', 3: 'online'}

        robot3.inRobots.extend([robot0, robot1, robot2])
        robot3.outRobots.extend([robot0, robot2])
        robot3.oldOutRobotsRecoder = {0: 'online', 2: 'online'}
        # endregion

        # region signal unit test

        # test robot.toOnlinePre()函数
        # stage.addNewRobots('signal')

        # test robot.toOnline()函数
        # stage.addNewRobots('signal')

        # endregion

        stage.addNewRobots('signal')

        stage.run()

        logging.info('------------------------------------END A NEW LOOP ------------------------------\n'
                     '------------------------------------END A NEW LOOP ------------------------------\n'
                     '------------------------------------END A NEW LOOP ------------------------------')

