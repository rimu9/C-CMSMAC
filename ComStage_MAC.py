# -*- coding: utf-8 -*-
"""
       .==.        .==.
      //`^\\      //^`\\
     // ^ ^\(\__/)/^ ^^\\
    //^ ^^ ^/+  0\ ^^ ^ \\
   //^ ^^ ^/( >< )\^ ^ ^ \\
  // ^^ ^/\| v''v |/\^ ^ ^\\
 // ^^/\/ /  `~~`  \ \/\^ ^\\
 ----------------------------
BE CAREFULL! THERE IS A DRAGON.

功能：待写
备注：

案例：

模块：
(c) HUANG JI 2019
依赖：
"""
import math
import sys

isCupy = False
import random
import gc

try:
    import cupy as np

    isCupy = True
except:
    import numpy as np

    isCupy = False
import copy
import matplotlib.pyplot as plt
from matplotlib import cbook
from matplotlib import cm
from matplotlib.colors import LightSource
import networkx as net
import openpyxl
from openpyxl import load_workbook
from common.DrKDtree import KDtree
from common.utils import *
import common.settings as settings
from Communication.simulation.ComObjectCollection import *
from Communication.simulation.ComObject import ComObject
from Communication.simulation.ComStage import ComStage
from Communication.simulation.ComRobotAF_MAC import ComRobotAF_MAC
from Communication.simulation.ComRobot import ComRobot


class ComStageMAC(ComStage):
    nowTime = 0  # 当前时间点记录

    def __init__(self):
        super().__init__()
        self.numRobot = 100  # 实验最大多少个机器人

        self.frameHZ = 10  # 一秒多少帧
        self.formerSlots = 20  # 一帧多少正式槽
        self.preSlots = 5  # 一帧多少预备槽

        self.rbtID_Time = []  # 这个数组里面，为每个slot定义一个数组，里面存放在该时刻发消息的机器人id
        for i in range(self.formerSlots + self.preSlots):
            self.rbtID_Time.append(set())

        '''
        SLOT-CSMA: slot base CSMA
        BASE-TDMA: base TDMA
        C-MSTDMA: cross-layer multi-slots TDMA
        C-CMSTDMA: cluster based cross-layer multi-slots TDMA
        '''
        self.macProtocol = 'base-TDMA'

        # 测试指标需要的变量
        self.acesSuccessFlag = 1  # 机器人成功接入网络，再生成新的机器人进来。
        self.accessCountHistory = []  # 储存每个机器人接入时间
        self.clusters = []  # 储存聚集系数
        self.shortestPathLen = []  # 储存平均最短路径
        self.overloadCount = 0  # 储存随着节点的增多，整体的接入负载。
        self.overloadHistory = []

        self.writexlsxFlag = 0
        self.returnFLag = 0
        self.addRobotType = 'multi'

        # 以下变量就是为了构建地图，对通信本身没啥用
        self.map = []
        self.leaveHistory = []
        self.linenum = 12
        self.scale = self.mEnvSize[0]

    def initMap(self):
        self.map = [[] for i in range(self.linenum - 2)]
        for i in range(1, self.linenum - 1):
            for j in range(1, self.linenum - 1):
                x = i / self.linenum * self.scale - self.scale / 2
                y = j / self.linenum * self.scale - self.scale / 2
                self.map[i - 1].append({'status': 'idle',  # idle空闲，waiting等待机器人到位，busy占用, useless该位置无法接入
                                        'pos': (x, y),
                                        'angle': None,
                                        'slots': set()})

    def refreshMap(self):
        for i in range(self.linenum - 2):
            for j in range(self.linenum - 2):
                if self.map[i][j]['status'] == 'useless':
                    self.map[i][j]['status'] = 'idle'
                    print('%%%%%%%%%%%%%%%%%%%%%%%%%%%%% 清除一个无用的位置', i, j, '%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%')

    def updateMap(self, robot: ComRobotAF_MAC):
        if robot.status == 'leaving':
            i = round((robot.mPos[0] + self.scale / 2) / self.scale * self.linenum - 1)
            j = round((robot.mPos[1] + self.scale / 2) / self.scale * self.linenum - 1)
            self.leaveHistory.append((i, j))
            self.map[i][j]['status'] = 'idle'
            self.map[i][j]['angle'] = robot.mDirection
            self.map[i][j]['slots'] = robot.slots
            self.leaveHistory.append((i, j))
            return
        if robot.status == 'entering':
            if robot.mPos[2] == 400:
                if len(self.leaveHistory) > 0:
                    i, j = self.leaveHistory.pop()
                    robot.mTarget = np.array([self.map[i][j]['pos'][0],
                                              self.map[i][j]['pos'][1],
                                              500], dtype=np.float32)
                    robot.p_acsInfo[robot.mId] = {'type': 'slot-rpl', 'slots': self.map[i][j]['slots']}
                    self.map[i][j]['status'] = 'waiting'
                    return
                else:  # 找一个位置让机器人过去，作为一个新机器人接入
                    for i in range(5):
                        mi = 4 - i
                        ma = 5 + i
                        for p in range(mi, ma + 1):
                            if self.map[mi][p]['status'] == 'idle':
                                self.map[mi][p]['status'] = 'waiting'
                                robot.mTarget = np.array([self.map[mi][p]['pos'][0],
                                                          self.map[mi][p]['pos'][1],
                                                          500], dtype=np.float32)
                                return
                        for p in range(mi, ma + 1):
                            if self.map[ma][p]['status'] == 'idle':
                                self.map[ma][p]['status'] = 'waiting'
                                robot.mTarget = np.array([self.map[ma][p]['pos'][0],
                                                          self.map[ma][p]['pos'][1],
                                                          500], dtype=np.float32)
                                return
                        for p in range(mi, ma + 1):
                            if self.map[p][mi]['status'] == 'idle':
                                self.map[p][mi]['status'] = 'waiting'
                                robot.mTarget = np.array([self.map[p][mi]['pos'][0],
                                                          self.map[p][mi]['pos'][1],
                                                          500], dtype=np.float32)
                                return
                        for p in range(mi, ma + 1):
                            if self.map[p][ma]['status'] == 'idle':
                                self.map[p][ma]['status'] = 'waiting'
                                robot.mTarget = np.array([self.map[p][ma]['pos'][0],
                                                          self.map[p][ma]['pos'][1],
                                                          500], dtype=np.float32)
                                return

                    # 遍历了所有的选择都不能找到一个好的位置接入，则随机找一个位置接入
                    range_x = (-700, 700)
                    range_y = (-700, 700)
                    x = random.uniform(range_x[0], range_x[1])
                    y = random.uniform(range_y[0], range_y[1])
                    robot.mTarget = np.array([x, y, 500], dtype=np.float32)
                    robot.randomInFlag = 1
                    print('______________ 找不到位置了，我随机接入了: id=', robot.mId, ' ————————————————————————————-——————')
            else:
                return

    def update(self):
        # 图坐标系建立
        if self.mAx:  # 画3d图的准备
            self.mAx.cla()
            self.mAx.grid(False)
            self.mAx.set_xlim(-self.mEnvSize[0], self.mEnvSize[0])
            self.mAx.set_ylim(-self.mEnvSize[1], self.mEnvSize[1])
            self.mAx.set_zlim(-self.mEnvSize[2], self.mEnvSize[2])
        if self.mGraphAx:  # 画连通图的准备
            self.mGraphAx.cla()

        # 梳理网络结构，确定有向连通图
        self.mRobotList.clear_edges()
        # kd_tree = KDtree(self.mRobotPosList)

        # 此处循环是为了感知周围的邻居，获得出邻居们
        for ind, pos in enumerate(self.mRobotPosList):
            robot = list(self.mRobotList.nodes)[ind]
            # if robot.status != 'entering':  # 对于正在进入的机器人，等其到达指定位置上了再进行感知。
            robot.refresh()
            if robot.isCommunicating:
                # 获得通信范围内的所有机器人,获得出邻居机器人
                robot.sense()

        # 在获得出邻居的基础上，增加判断，获得入邻居，但是这里会将有冲突的机器人也加入到如邻居中来，需要进一步判断
        for ind, pos in enumerate(self.mRobotPosList):
            robot = list(self.mRobotList.nodes)[ind]
            if robot.isCommunicating:
                for rbt in robot.robots_inRs:
                    if rbt.status == 'entering':
                        continue
                    if robot in rbt.out_robots:
                        robot.in_robots.append(rbt)  # 获得入邻居机器人

        # 此处的循环是每个机器人通过互斥表信息判断自己是否需要舍弃某些slot
        tempSet = set()
        for ind, pos in enumerate(self.mRobotPosList):
            robot = list(self.mRobotList.nodes)[ind]
            if robot.isCommunicating:
                tempSet |= robot.inCollisionDetect()
        if tempSet:
            print('///////////////////////for some robot should remove a slot: ', tempSet)

        # 此处就是通过judgeConflict函数进行判断是否有冲突存在，并在图中绘制不同颜色的连接线。
        edges_color = []
        for ind, pos in enumerate(self.mRobotPosList):
            robot = list(self.mRobotList.nodes)[ind]
            if robot.isCommunicating:
                # 连通图内各个机器人通信动作冲突性判断
                robot.in_robots, rept_robot = self.judgeConflict(robot.in_robots)
                for rpt_rbt in rept_robot:
                    rpt_rbt.out_robots.remove(robot)
                    self.mRobotList.add_edge(rpt_rbt, robot)
                    edges_color.append('r')
                if len(rept_robot) > 0:
                    print('!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!')
                    print('out:', [rbt.mId for rbt in rept_robot], ' in:', robot.mId)
                    print('id:', robot.mId, 'in neighbors: ', [rbt.mId for rbt in robot.in_robots])
                    for r in rept_robot:
                        print('id:', r.mId, 'out neighbors: ', [rbt.mId for rbt in r.out_robots])
                    print('!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!')
                for rbt in robot.in_robots:
                    self.mRobotList.add_edge(rbt, robot)
                    edges_color.append('b')

        # 执行正确通信动作
        print('nowTime: ', ComStageMAC.nowTime, ' with id:', self.rbtID_Time[ComStageMAC.nowTime], 'can send msg.')
        for vrobot in self.mRobotList.nodes:
            for rbt in vrobot.out_robots:
                self.overloadCount += vrobot.communicateTo(rbt, ComStageMAC.nowTime)
            if vrobot.isCommunicating and (ComStageMAC.nowTime in vrobot.slots):
                vrobot.sendMsg.clear()  # 发送完成后清空
                vrobot.hasSendLeaFlag = 2  # 对于要离开的机器人来说，直到这时才将离开洪泛包发出去，然后机器人update时就可以离开网络了

        # 执行完通信操作以后，就把tempSet中的数据重新加回来，因为每次机器人接收的信息可能不同步，不是每次判断都是对的，所以采取的措施就是我每次根据互斥信息进行判断，而不是判断一次就永久删除。
        for item in tempSet:
            robot = list(self.mRobotList.nodes)[item[0]]
            if robot.mId != item[0]:
                print('ERROR ERROR ERROR !!! ERROR: Line 254, ')
                return
            robot.slots.add(item[1])

        # 执行机器人自身的更新程序，融合数据
        for ind, vrobot in enumerate(self.mRobotList.nodes):
            # 如果机器人没有目标位置，则stage分配一个
            self.updateMap(vrobot)
            # 机器人自身的状态更新
            vrobot.update()
            # 将新的机器人位置信息填入到self.mRobotPosList中
            self.mRobotPosList[ind] = vrobot.mPos
            # 如果需要，在图中画出机器人图像，会引起程序运行缓慢。
            if self.mAx:
                vrobot.draw(self.mAx)

        # 图形显示
        if self.isPlotGraph and len(self.mRobotList.nodes) > 1:
            graph_pos = net.circular_layout(self.mRobotList)  # 布置框架
            net.draw(self.mRobotList, graph_pos, ax=self.mGraphAx,
                     with_labels=False, node_size=30, edge_color=edges_color)

        # 初始化设置的4个机器人交互需要一定的时间才能使初始化网络稳定下来，设定35个frame后新机器人才能开始接入。
        if self.count == 35:
            self.acesSuccessFlag = 0

        if self.acesSuccessFlag == 0:
            # 清除map中设置为无用的点
            self.refreshMap()
            # 计算聚集系数等指标
            self.calcNetworkMetrics()
            # 记录网络负载
            num = len(self.mRobotList)
            self.overloadHistory.append((num, self.overloadCount))
            # 新生成一个机器人或者多个机器人
            num = self.addNewRobots(self.addRobotType)

            self.acesSuccessFlag = num

        # 更新时间点
        ComStageMAC.nowTime += 1
        if ComStageMAC.nowTime >= (self.formerSlots + self.preSlots):
            ComStageMAC.nowTime = 0

        # 设置当网络有50个机器人的时候，就停止仿真，毕竟是小尺寸网络， 后期可以试试大一点的网络。
        if len(self.mRobotList) == 50 and self.writexlsxFlag == 0:
            self.writeLogFile()
            self.writexlsxFlag = 1
            self.returnFLag = 1

    def judgeConflict(self, in_rbt):
        if len(in_rbt) == 0:
            return [], []
        rept_rbt = []
        in_rbt_temp = []
        in_rbt_id = [robot.mId for robot in in_rbt]
        rept_set = set(in_rbt_id) & set(self.rbtID_Time[ComStageMAC.nowTime])
        if len(rept_set) > 1:  # 大于1表示有多个机器人在该时刻同时发给我
            # 记录冲突的机器人
            for rbt in in_rbt:
                if rbt.mId in rept_set:
                    rept_rbt.append(rbt)
                else:
                    in_rbt_temp.append(rbt)
        else:
            return in_rbt, []

        return in_rbt_temp, rept_rbt

    def calcNetworkMetrics(self):
        num = len(self.mRobotList)
        cluster = net.average_clustering(self.mRobotList)
        self.clusters.append((num, cluster))
        # print('robots num: ', num, ' with clusters: ', cluster)
        shortPath = net.average_shortest_path_length(self.mRobotList)
        self.shortestPathLen.append((num, shortPath))
        # print('robots num: ', num, ' with shortest path: ', shortpath)

    def addNewRobots(self, addType):
        if addType == 'signal':
            num = 1
            range_x = (-stage.mEnvSize[0], stage.mEnvSize[0])
            range_y = (-stage.mEnvSize[1], stage.mEnvSize[1])
            x = random.uniform(range_x[0], range_x[1])
            y = random.uniform(range_y[0], range_y[1])
            newrobot = ComRobotAF_MAC((x, y, 400), stage)
            newrobot.isPlotTrail = True
            newrobot.setRadius(2)
            stage.addRobot(newrobot)
            print('**********a new robot has generate ---id:', newrobot.mId, '*****************')
        else:
            num = random.randint(2, 4)
            for i in range(num):
                range_x = (-stage.mEnvSize[0], stage.mEnvSize[0])
                range_y = (-stage.mEnvSize[1], stage.mEnvSize[1])
                x = random.uniform(range_x[0], range_x[1])
                y = random.uniform(range_y[0], range_y[1])
                newrobot = ComRobotAF_MAC((x, y, 400), stage)
                newrobot.isPlotTrail = True
                newrobot.setRadius(2)
                stage.addRobot(newrobot)
                print('**********a new robot has generate ---id:', newrobot.mId, '*****************')

        return num

    def writeLogFile(self):
        file1 = load_workbook("../../resultHJ/accessTime_200.xlsx")
        sheet1 = file1[file1.sheetnames[0]]
        max1 = sheet1.max_row + 1
        l = 1
        for item in self.accessCountHistory:
            sheet1.cell(max1, l, str(item))
            l += 1
        file1.save("../../resultHJ/accessTime_200.xlsx")

        file2 = load_workbook("../../resultHJ/clusters_200.xlsx")
        sheet2 = file2[file2.sheetnames[0]]
        max2 = sheet2.max_row + 1
        l = 1
        for item in self.clusters:
            sheet2.cell(max2, l, str(item))
            l += 1
        file2.save("../../resultHJ/clusters_200.xlsx")

        file3 = load_workbook("../../resultHJ/shortestpath_200.xlsx")
        sheet3 = file3[file3.sheetnames[0]]
        max3 = sheet3.max_row + 1
        l = 1
        for item in self.shortestPathLen:
            sheet3.cell(max3, l, str(item))
            l += 1
        file3.save("../../resultHJ/shortestpath_200.xlsx")

        file4 = load_workbook("../../resultHJ/overload_200.xlsx")
        sheet4 = file4[file4.sheetnames[0]]
        max4 = sheet4.max_row + 1
        l = 1
        for item in self.overloadHistory:
            sheet4.cell(max4, l, str(item))
            l += 1
        file4.save("../../resultHJ/overload_200.xlsx")

        print('ACCESS TIME: ', self.accessCountHistory)
        print('CLUSTERS: ', self.clusters)
        print('SHORTEST PATH LEN: ', self.shortestPathLen)
        print('OVERLOAD: ', self.overloadHistory)

    def run(self):
        # super().run()
        self.initEnv()
        while True:
            if self.count % self.mSavePosRound == 0:
                if self.isSavePos2:
                    self.isSavePos = True
            else:
                self.isSavePos = False
            if (self.count * settings.CS_INTERVAL) >= self.mRuningTime:
                break
            self.count += 1
            # if self.count % 1 == 0:
            #     print("Round: 【%d : %d】" % (self.count, self.mRuningTime/settings.CS_INTERVAL))
            self.update()
            plt.pause(settings.CS_INTERVAL)

            if self.returnFLag == 1:
                print('程序运行一次退出')
                break


if __name__ == "__main__":
    initdir = copy.deepcopy(list(globals().keys()))
    for i in range(50):
        for key in list(globals().keys()):
            if (key not in initdir) and (key != "key") and (key != "initdir") and (key != "i"):
                globals().pop(key)
        del key
        object_collection.clear()
        gc.collect()

        ComRobot._robot_count = 0
        ComStage.mCount = 0
        ComStageMAC.nowTime = 0

        stage = ComStageMAC()
        stage.initMap()
        # 初始化4个机器人作为原始机器人
        robot0 = ComRobotAF_MAC((stage.map[4][4]['pos'][0], stage.map[4][4]['pos'][1], 500), stage)
        stage.map[4][4]['status'] = 'busy'
        robot0.setDirection(60 / 180 * math.pi)
        robot1 = ComRobotAF_MAC((stage.map[4][5]['pos'][0], stage.map[4][5]['pos'][1], 500), stage)
        stage.map[4][5]['status'] = 'busy'
        robot1.setDirection(-40 / 180 * math.pi)
        robot2 = ComRobotAF_MAC((stage.map[5][4]['pos'][0], stage.map[5][4]['pos'][1], 500), stage)
        stage.map[5][4]['status'] = 'busy'
        robot2.setDirection(135 / 180 * math.pi)
        robot3 = ComRobotAF_MAC((stage.map[5][5]['pos'][0], stage.map[5][5]['pos'][1], 500), stage)
        stage.map[5][5]['status'] = 'busy'
        robot3.setDirection(-100 / 180 * math.pi)
        robot0.slots.add(0)
        robot0.slots.add(1)
        robot1.slots.add(2)
        robot1.slots.add(3)
        robot2.slots.add(4)
        robot2.slots.add(5)
        robot3.slots.add(6)
        robot3.slots.add(7)
        robotlist = [robot0, robot1, robot2, robot3]
        for robot in robotlist:
            robot.setRadius(2)
            robot.status = 'online'
            robot.resFlag = 1
            robot.hasSlotChangedFlag = 1
            robot.p_acsInfo[robot.mId]['slots'] = copy.deepcopy(robot.slots)
            for s in robot.slots:
                stage.rbtID_Time[s].add(robot.mId)
        stage.addRobot(robot0)
        stage.addRobot(robot1)
        stage.addRobot(robot2)
        stage.addRobot(robot3)

        # for i in range(2):
        #     range_x = (-stage.mEnvSize[0], stage.mEnvSize[0])
        #     range_y = (-stage.mEnvSize[1], stage.mEnvSize[1])
        #     range_z = (-stage.mEnvSize[2], stage.mEnvSize[2])
        #     x = random.uniform(range_x[0], range_x[1])
        #     y = random.uniform(range_y[0], range_y[1])
        #     z = random.uniform(range_z[0], range_z[1])
        #     robot = ComRobotAF_MAC((x, y, 400), stage)
        #     # robot.mTarget = np.array([s5[i], s6[i], 500], dtype=np.float32)
        #     robot.isPlotTrail = True
        #     stage.addRobot(robot)

        # robot = ComRobotAF_MAC((0, 0, 400), stage)
        # robot.setDirection(-175 / 180 * math.pi)
        # robot1 = ComRobotAF_MAC((-10, 100, 400), stage)
        # robot.setRadius(2)
        # robot1.setRadius(2)
        # stage.addRobot(robot)
        # stage.addRobot(robot1)

        stage.run()


        # 接入时候打印一下出入邻居情况，出现机器人一直不能接入的情况，占用pre-slot。
        # 调整机器人占用pre-slot策略，出现机器人同事占用一个pre-slot都无法接入的情况。 sol：随机换着接入。
        # 接收到同一个机器人的请求，忽略后面的请求  sol：增加记录数组，判断。
        # 接入碰撞消息没有起作用，还是存在接入碰撞-----新机器人的出邻居中有其他新接入机器人b，那么b需不需要response？--sol:机器人接入后只有出邻居中是准备或者上线状态情况下，才发送请求消息，机器人b不response
        # response动作的chooseSlot函数中，通过自身对网络信息的了解来尽可能去重的计算过程中，没有将自己剔除，导致网络一直在发送自己与自己存在冲突，网络不能正常运行  sol：函数中增加剔除自身的代码。
        # 机器人发现useless位置后随机选取的初始化位置太远了，时间长。  sol：随机值选取范围选择小一点。，了，